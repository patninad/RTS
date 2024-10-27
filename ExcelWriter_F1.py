from ExcelWriter import ExcelWriter
from collections import OrderedDict
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image
import json
import pathlib

def equalise_lists_zip(l1, l2, default=""):
    # Add "" to smaller info array until they are the same size for zipping
    len_1 = len(l1)
    len_2 = len(l2)
    if len_1 != len_2:
        if min(len_1, len_2) == len_1:
            min_list = l1
        elif min(len_1, len_2) == len_2:
            min_list = l2

        for _ in range(abs(len_2 - len_1)):
            min_list.append(default) 

class ExcelWriter_F1(ExcelWriter):
    def __init__(self, converter, amount_percentage, incl_gst):
        super().__init__(converter, amount_percentage, incl_gst)

        # Import json config files
        self.template_info_json = pathlib.Path.cwd().joinpath("format_info", "template_info_f1.json")
        self.template_info = json.load(open(self.template_info_json), object_pairs_hook=OrderedDict)

        self.person_info_json = pathlib.Path.cwd().joinpath("format_info", "Acc", f"{self.converter.person}.json")
        self.person_info = json.load(open(self.person_info_json))

        self.general_info_json = pathlib.Path.cwd().joinpath("format_info", f"general_f1.json")
        self.general_info = json.load(open(self.general_info_json))

        # Update person_info with general_info
        self.general_info.update(self.person_info)
        self.person_info = self.general_info.copy()

        self.font = Font(name="Arial", size="8")

    def write_wb(self):
        self.write_intro()
        self.write_header()
        self.write_data_rows()

        self.write_total_info()

        self.add_page_break()

        self.write_deductions_info()
        self.write_summary_info()

    def write_intro(self):
        # Introduction
        for info_row in self.template_info["intro_text_header"]:
            self.ws.merge_cells(f"D{self.current_row}:G{self.current_row}")
            self.ws[f"D{self.current_row}"] = info_row
            self.ws[f"D{self.current_row}"].alignment = Alignment(horizontal="center")
            self.ws[f"D{self.current_row}"].font = Font(name=self.font.name, size="10", bold=True)
            self.goto_next_row()
        
        # Invoice Info
        count = 1
        for invoice_info_row in self.converter.invoice_info:
            self.ws[f"H{count}"] = invoice_info_row
            self.ws[f"H{count}"].font = Font(name=self.font.name, size="10")
            self.ws[f"K{count}"] = self.converter.invoice_info[invoice_info_row]
            self.ws[f"K{count}"].alignment = Alignment(horizontal="right")
            count += 1

        self.add_empty_lines(2)

        # Add border images
        img = Image(pathlib.Path.cwd().joinpath("format_info", "intro_border.gif"))
        img.height = 150
        img.width = 330
        img.anchor = f"A{self.current_row - 1}"
        self.ws.add_image(img)
        img2 = Image(pathlib.Path.cwd().joinpath("format_info", "intro_border.gif"))
        img2.height = 150
        img2.width = 330
        img2.anchor = f"F{self.current_row - 1}"
        self.ws.add_image(img2)

        # Person (from) and (to) data 
        self.ws[f"A{self.current_row}"] = "From: "
        self.ws[f"A{self.current_row}"].font = Font(name=self.font.name, size="10", bold=True)
        self.ws[f"A{self.current_row}"].alignment = Alignment(horizontal="right")

        self.ws[f"F{self.current_row}"] = "To: "
        self.ws[f"F{self.current_row}"].alignment = Alignment(horizontal="right")
        self.ws[f"F{self.current_row}"].font = Font(name=self.font.name, size="10", bold=True)

        # Add "" to smaller info array until they are the same size for zipping
        equalise_lists_zip(self.person_info["data"], self.template_info["intro_text_right"])

        for person_row, intro_row in zip(self.person_info["data"], self.template_info["intro_text_right"]):
            self.ws[f"B{self.current_row}"] = f"    {person_row}"
            self.ws[f"G{self.current_row}"] = f"    {intro_row}"
            self.goto_next_row()

        self.add_empty_lines(2)
        self.ws[f"A{self.current_row}"] = "Job Details"
        self.ws[f"A{self.current_row}"].font = Font(name=self.font.name, size="8", bold=True, underline="single")
        self.add_empty_line()
        
        self.goto_next_row()     

    # Write column header
    def write_header(self):
        self.write_list_to_row(self.current_row, self.columns, bold_l=True)
        self.goto_next_row() 
    
    # Write data rows to wb (i.e., rows with specific column values in self.column_values)
    # Record the first row numbers with new date values
    def write_data_rows(self):
        new_date_rows = []
        subtotal_rows = []
        prev_date = None

        for row_dict in self.column_values:

            # Stop when it reaches the total section
            if "Sub" in row_dict["Service"] and "Total" in row_dict["Service"] \
                or "Sub" in row_dict["Description"] and "Total" in row_dict["Description"]:
                break

            # If no date has been encountered yet or a new date is encountered
            if row_dict["Date"] and (not prev_date or prev_date != row_dict["Date"]):
                prev_date = row_dict["Date"]
                new_date_rows.append(self.current_row)

            # If found a subtotal row
            if row_dict["Date"] and row_dict["Service"] == "Total":
                subtotal_rows.append(self.current_row)
                
            # If rate includes one of 'tolls', don't write row to excel file
            if row_dict["Service"] in self.person_info["tolls"]:
                self.stats["total_tolls_removed"] += 1
                self.stats["total_tolls_amount"] += float(row_dict["Amount $"])

                if row_dict["Service"] in self.person_info["tolls_add"]:
                    add_toll_entry = f'''{row_dict["Service"]}_tolls_removed'''
                    self.stats[add_toll_entry] = self.stats[add_toll_entry] + 1 if add_toll_entry in self.stats else 1
            else:
                self.write_dict_to_row(self.current_row, row_dict)
                self.goto_next_row()

        # Write subtotals
        self.write_subtotals(new_date_rows, subtotal_rows)

    def write_dict_to_row(self, row, row_dict):
        # Write row values using dict with column names as keys 
        for col, value in row_dict.items():
            coord = f"{self.col_map[col]}{row}"

            # Cast numeric values to float
            try:
                value = float(value)
            except ValueError:
                value = value.encode("utf-8")

            # Write a percent of the original amount price
            if value and col == "Amount $":
                value = self.amount_percentage * value

            self.ws[coord] = value
            self.apply_formatting(coord, col)

    def write_total_info(self):
        # Remove one empty line

         # Check for GST inclusion
        if self.incl_gst == "n":
            # If gst is not included
            subtotals_key = "subtotals_total_info_without_GST"
            charges_key = "Sub-Total"
        else:
            subtotals_key = "subtotals_total_info_with_GST"
            charges_key = "Sub-Total (Incl. GST)"

        self.template_info[subtotals_key][charges_key] = f'''=SUM({"+".join(self.stats["subtotal_coords"])})'''

        added_tolls_amount = sum([self.stats[f'''{toll}_tolls_removed'''] for toll in self.person_info["tolls_add"]])
        self.template_info[subtotals_key]["Tolls (Incl. GST)"] = f'''={self.toll_price} * {added_tolls_amount}'''

        self.write_section(subtotals_key)
        
    def write_deductions_info(self):
        # keys in self.template_info
        deductions_sections = ["deductions_info", "reliable_t_deduction_info"]

        for deduction_section in deductions_sections:
            self.write_section(deduction_section)

    def write_subtotals(self, new_date_rows, subtotal_rows):
        date_index = 0
        self.stats["subtotal_coords"] = []

        for subtotal_row in subtotal_rows:
            prev_date_row = new_date_rows[date_index]

            # Shift Date to merged cell
            self.ws.merge_cells(f'''{self.col_map["Kms"]}{subtotal_row}:{self.col_map["UOM"]}{subtotal_row}''')
            self.ws[f'''{self.col_map["Kms"]}{subtotal_row}'''] = self.ws[f'''{self.col_map["Date"]}{subtotal_row}'''].value

            self.ws[f'''{self.col_map["Kms"]}{subtotal_row}'''].alignment = Alignment(horizontal="center")
            self.ws[f'''{self.col_map["Kms"]}{subtotal_row}'''].border = Border(top=Side(style="thin"))
            self.ws[f'''{self.col_map["Qty"]}{subtotal_row}'''].border = Border(top=Side(style="thin"))
            self.ws[f'''{self.col_map["UOM"]}{subtotal_row}'''].border = Border(top=Side(style="thin"))

            self.ws[f'''{self.col_map["Date"]}{subtotal_row}'''] = ""

            # Shift service to description
            self.ws[f'''{self.col_map["Description"]}{subtotal_row}'''] = self.ws[f'''{self.col_map["Service"]}{subtotal_row}'''].value + ":"
            self.ws[f'''{self.col_map["Description"]}{subtotal_row}'''].alignment = Alignment(horizontal="right")
            self.ws[f'''{self.col_map["Service"]}{subtotal_row}'''] = ""

            amount_col = self.col_map["Amount $"]
            # Subtotal from previous date amount to row before subtotal
            subtotal_val = f'''=SUM({amount_col}{prev_date_row}:{amount_col}{subtotal_row - 1})'''

            # Add all values from last date to before subtotal
            for row in range(prev_date_row, subtotal_row):
                amount_val = self.ws[f'''{self.col_map["Amount $"]}{row}'''].value
                self.amounts.append(amount_val)

                try:
                    self.stats["total_amount"] += float(amount_val)
                except ValueError:
                    pass

            coord = f'''{amount_col}{subtotal_row}'''
            self.ws[coord] = subtotal_val
            self.apply_formatting(coord, amount_col, bold=True)
            self.ws[coord].border = Border(top=Side(style="thin"))

            self.stats["subtotal_coords"].append(f'''{amount_col}{subtotal_row}''')

            date_index += 1

    def write_summary_info(self):
        self.template_info["summary_info"]["Amount Payable on Recipient Created Tax Invoices"] = f'''={self.stats["total_cells"][0]}'''
        self.template_info["summary_info"]["Less - TRANZWORKS PTY LTD"] = f'''={self.stats["total_cells"][1]}'''
        self.template_info["summary_info"]["Less - RELIABLE TRANS SERV"] = f'''={self.stats["total_cells"][2]}'''

        self.write_section("summary_info")
    
    def write_section(self, section_key, extra_space=False):
        self.add_empty_lines(1)

        space = not extra_space

        if self.template_info[section_key].items():
            self.write_section_heading(section_key, space=space)

        total_start_row = self.current_row if not extra_space else self.current_row + 1
        gst_start_row = self.current_row if not extra_space else self.current_row + 1

        toll_row = ""
        invoice_excl_row = ""
        gst_row = ""

        for info_row, val in self.template_info[section_key].items():
            if extra_space:
                self.add_empty_line()

            # Write text
            coord = f'''{self.col_map["To"]}{self.current_row}'''
            self.ws[coord] = info_row
            self.apply_formatting(coord, self.columns[1], ignore_num_f=True, ignore_align=True, bold=True)
        
            # Write value
            coord = f'''{self.col_map["Amount $"]}{self.current_row}'''
            
            if info_row in ["GST", "Add GST"]:
                val = f'''=0.1*SUM({self.col_map["Amount $"]}{gst_start_row}:{self.col_map["Amount $"]}{self.current_row - 1})'''
                gst_row = self.current_row
                self.apply_formatting(coord, "Amount $")

            elif info_row == "Tolls (Incl. GST)":
                # Can't include toll in total sums so reinitialise row that starts the section
                self.apply_formatting(coord, "Amount $")
                # Add toll cell to totals information
                toll_row = self.current_row

            elif info_row == "Sub-Total":
                total_start_row = self.current_row
                self.apply_formatting(coord, "Amount $")

                if not self.subtotal_row:
                    self.subtotal_row = total_start_row

            elif info_row == "Invoice Excl. GST":
                val = f'''=SUM({self.col_map["Amount $"]}{total_start_row}:{self.col_map["Amount $"]}{self.current_row - 1})'''
                total_start_row = self.current_row
                gst_start_row = self.current_row
                invoice_excl_row = self.current_row
                self.apply_formatting(coord, "Amount $")

            elif info_row == "Invoice Incl. GST" in info_row:
                if toll_row and invoice_excl_row:
                    val = f'''={self.col_map["Amount $"]}{toll_row} + {self.col_map["Amount $"]}{invoice_excl_row} + {self.col_map["Amount $"]}{gst_row}'''
                else:
                    val = f'''=SUM({self.col_map["Amount $"]}{total_start_row}:{self.col_map["Amount $"]}{self.current_row - 1})'''
                self.apply_formatting(coord, "Amount $")

                self.stats["total_cells"].append(coord)

            elif info_row == "NET AMOUNT DUE":
                total_payable_cell = f'''{self.col_map["Amount $"]}{total_start_row}'''
                total_deduction_range = f'''{self.col_map["Amount $"]}{total_start_row + 1}:{self.col_map["Amount $"]}{self.current_row - 1}'''
                val = f'''= {total_payable_cell} - SUM({total_deduction_range})'''
                self.apply_formatting(coord, "Amount $", bold=True)

            else:
                self.apply_formatting(coord, "Amount $")

            self.ws[coord] = val
            self.goto_next_row()
            self.add_empty_line()

    def write_section_heading(self, key, space=True):
        possible_header = key.replace("_info", "") + "_header"

        # Write section header, if it exists in self.template_info
        if possible_header in self.template_info:
            coord = f"B{self.current_row}"
            self.ws[coord] = self.template_info[possible_header]
            self.ws[coord].font = Font(name=self.font.name, size="10", bold=True, underline="single")
            # self.apply_formatting(coord, self.columns[1], ignore_num_f=True, ignore_align=True)

            self.goto_next_row()
            
            if space:
                self.add_empty_line()

    # Set custom widths for excel file columns
    def set_col_widths(self):
        if "Description" in self.columns:
            col_widths = { "Date": 7.55,
                        "Job No.": 5.73,
                        "Client": 9.91,
                        "From": 11.18,
                        "To": 11.45,
                        "Service": 6,
                        "Description": 15.36,
                        "Kms": 3.82,
                        "Qty": 4.73,
                        "UOM": 4.18,
                        "Amount $": 10.18,
                        }
        else:
            col_widths = { "Date": 14.12,
                        "Job No.": 10.18,
                        "From": 13.27,
                        "To": 13.36,
                        "Rate": 12.73,
                        "Amount $": 16
                        }

        for col in col_widths:
            self.ws.column_dimensions[self.col_map[col]].width = col_widths[col] + 0.6

    def apply_formatting(self, coord, col, ignore_num_f=False, ignore_align=False, bold=False):
        num_format = {
                        "Job No.": numbers.FORMAT_NUMBER,
                        "Amount $": '''_-$* #,##0.000_-;-$* #,##0.000_-;_-$* "-"??_-;_-@_-''',
                        "Qty": '''0.00'''
                     }
        align_format = {
                        # "Job No.": Alignment(horizontal="right"),
                        "Kms": Alignment(horizontal="right"),
                        "Qty": Alignment(horizontal="right"),
                        "Amount $": Alignment(horizontal="right"),
                       }
        default_align = Alignment(horizontal="left")

        # Apply number format and alignment formatting if row is not part of header
        if not ignore_num_f and col in num_format:
            self.ws[coord].number_format = num_format[col]

        if not ignore_align:
            self.ws[coord].alignment = align_format[col] if col in align_format else default_align
    
        self.ws[coord].font = Font(name=self.font.name, size=self.font.size, bold=True) if bold else self.font