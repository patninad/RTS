from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.worksheet.pagebreak import Break
from collections import OrderedDict
from Converter import check_create_folder
import pathlib
import json


def extract_person_info(person_file):
    with open(person_file, "r") as obj:
        lines = obj.read().split("\n")
        lines = [line.strip().title() if line[0].islower() else line.strip()
                 for line in lines]
    return {"data": lines}


class ExcelWriter:

    def __init__(self, converter, amount_percentage, incl_gst):
        # Converter object
        self.converter = converter
        self.columns = converter.get_columns()
        self.column_values = converter.get_column_values()
        self.incl_gst = incl_gst

        # Import json config files
        self.template_info_json = pathlib.Path.cwd().joinpath(
            "format_info", "template_info_f2.json")
        self.template_info = json.load(
            open(self.template_info_json), object_pairs_hook=OrderedDict)

        self.person_info_txt = pathlib.Path.cwd().parents[0].joinpath(
            "Acc", f"{self.converter.person}.txt")
        self.person_info = extract_person_info(self.person_info_txt)

        self.general_info_json = pathlib.Path.cwd().joinpath(
            "format_info", f"general_f2.json")
        self.general_info = json.load(open(self.general_info_json))

        # Update person_info with general_info
        self.general_info.update(self.person_info)
        self.person_info = self.general_info.copy()

        self.main_sheet_title = "Sheet"
        self.__font = Font(name="Courier New", size="11")

        self.amount_percentage = amount_percentage / 100
        self.toll_price = 10.2
        self.__toll_check_col = "Rate"

        self.current_row = 1

        self.stats = OrderedDict({"dates": [],
                                  "subtotal_coords": [],
                                  "after_date_toll_coords": [],
                                  "total_amount": 0,
                                  "total_amounts_by_date": {},
                                  "total_tolls_removed_by_date": {},
                                  "total_tolls_removed": 0,
                                  "total_tolls_amount_by_date": {},
                                  "total_tolls_amount": 0,
                                  "total_jobs_by_date": {},
                                  "total_cells": [],
                                  "total_cells_summary": []
                                  })

        self.amounts = []

        self.__subtotal_row = None

        # Create mappings from col names to excel column values (e.g. 'Date' -> 'A')
        self.create_col_map()

    # Create mapping from column names to excel column names (A, B, C, etc.)
    def create_col_map(self):
        cols = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z".split(",")
        cols = cols[:len(self.columns)]
        self.col_map = dict(zip(self.columns, cols))

    # Configure page settings (size and margins)
    def config(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.main_sheet_title

        self.page_setting()

    def page_setting(self):
        # Page settings
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LETTER
        self.ws.page_margins.left = 0.3
        self.ws.page_margins.right = 0.27
        self.ws.page_margins.bottom = 0.5
        self.ws.page_margins.top = 0.3

        self.set_col_widths()

    def write_wb(self):
        self.write_intro()
        self.write_header()
        self.write_data_rows()

        self.write_total_info()
        self.write_deductions_info()
        self.write_remittance_info()

    def write_summary(self):
        self.ws = self.wb.create_sheet()
        self.ws.title = "Summary"
        self.current_row = 1
        self.columns = ["Date", "Jobs", "Amount", "Cash Held"]
        self.amount_percentage = 1
        self.create_col_map()
        self.page_setting()

        # Actual Writing
        self.write_intro_summary()
        self.write_header()

        dates = self.stats["dates"]
        jobs = list(
            map(lambda date: self.stats["total_jobs_by_date"][date], self.stats["dates"]))
        amounts = list(map(
            lambda coord: f"={self.main_sheet_title}!{coord}", self.stats["subtotal_coords"]))
        cash_held = [""] * len(dates)

        self.column_values = list(map(lambda row: dict(
            zip(self.columns, row)), zip(dates, jobs, amounts, cash_held)))

        first_total_coord = f'''{self.col_map["Amount"]}{self.current_row}'''
        last_total_coord = ""
        for row_dict in self.column_values:
            self.write_dict_to_row(self.current_row, row_dict)
            last_total_coord = f'''{self.col_map["Amount"]}{self.current_row}'''
            self.goto_next_row()

        self.goto_next_row()

        self.write_total_info_summary(
            f'''=SUM({first_total_coord}: {last_total_coord})''')
        self.write_deductions_info(deductions_sections=["goods_services_info"])
        self.write_remittance_info_summary()

    def write_intro(self):
        # Introduction
        for info_row in self.template_info["intro_text"]:
            self.ws.merge_cells(f"C{self.current_row}:F{self.current_row}")
            self.ws[f"C{self.current_row}"] = info_row
            self.ws[f"C{self.current_row}"].alignment = Alignment(
                horizontal="center")
            self.ws[f"C{self.current_row}"].font = Font(
                name="Courier New", size="14")
            self.goto_next_row()

        # Person data
        self.add_empty_line()
        for data_row in self.person_info["data"]:
            self.ws[f"A{self.current_row}"] = data_row
            self.ws[f"A{self.current_row}"].font = self.font
            self.goto_next_row()

        # Invoice data
        self.add_empty_lines(2)
        for invoice_info_row in self.converter.invoice_info:
            self.ws[f"A{self.current_row}"] = invoice_info_row + \
                " " + self.converter.invoice_info[invoice_info_row]
            self.ws[f"A{self.current_row}"].font = self.font
            self.goto_next_row()

        self.goto_next_row()

    def write_intro_summary(self):
        # Introduction
        self.template_info["intro_text"] = [
            f"      {info_row}" for info_row in self.template_info["intro_text"]]
        for info_row in self.template_info["intro_text"]:
            self.ws.merge_cells(f"A{self.current_row}:D{self.current_row}")
            self.ws[f"A{self.current_row}"] = info_row
            self.ws[f"A{self.current_row}"].alignment = Alignment(
                horizontal="center")
            self.ws[f"A{self.current_row}"].font = Font(
                name="Courier New", size="14")
            self.goto_next_row()

        # Person data
        self.add_empty_line()
        for data_row in self.person_info["data"]:
            self.ws[f"A{self.current_row}"] = data_row
            self.ws[f"A{self.current_row}"].font = self.font
            self.goto_next_row()

        # Invoice data
        self.add_empty_lines(2)
        for invoice_info_row in self.converter.invoice_info:
            self.ws[f"A{self.current_row}"] = invoice_info_row + \
                " " + self.converter.invoice_info[invoice_info_row]
            self.ws[f"A{self.current_row}"].font = self.font
            self.goto_next_row()

        self.goto_next_row()

    def write_total_info(self):
        # Remove one empty line
        self.goto_prev_row(decrement=2)

        # Check for GST inclusion
        if self.incl_gst == "n":
            # If gst is not included
            subtotals_key = "subtotals_total_info_without_GST"
            charges_key = "Total freight charges"
        else:
            subtotals_key = "subtotals_total_info_with_GST"
            charges_key = "Total freight charges (Incl. GST)"

        self.template_info[subtotals_key][
            charges_key] = f'''={"+".join(self.stats["subtotal_coords"])}'''

        added_tolls_amount = f'''={"+".join(self.stats["after_date_toll_coords"])}'''
        self.template_info[subtotals_key]["Total Toll (Incl. GST)"] = added_tolls_amount

        self.write_section(subtotals_key)

    def write_total_info_summary(self, formula):
        # Remove one empty line
        self.goto_prev_row(decrement=2)

        # Check for GST inclusion
        if self.incl_gst == "n":
            # If gst is not included
            subtotals_key = "subtotals_total_info_without_GST"
            charges_key = "Total freight charges"
        else:
            subtotals_key = "subtotals_total_info_with_GST"
            charges_key = "Total freight charges (Incl. GST)"

        self.template_info[subtotals_key][charges_key] = formula

        added_tolls_amount = f"={self.main_sheet_title}!{self.total_toll_coord}"
        self.template_info[subtotals_key]["Total Toll (Incl. GST)"] = added_tolls_amount

        self.write_section(subtotals_key, source='summary')

    def write_deductions_info(self, deductions_sections=[]):
        # keys in self.template_info
        if not deductions_sections:
            deductions_sections = ["goods_services_info", "rcit_deduction_info",
                                   "reliable_t_deduction_info", "reliable_t_mob_deduction_info"]

        # Add extracted goods info in template
        self.converter.goods_info.update(
            self.template_info["goods_services_info"])
        self.template_info["goods_services_info"] = self.converter.goods_info.copy(
        )

        for deduction_section in deductions_sections:
            self.write_section(deduction_section)

    def write_remittance_info(self):
        print(self.stats)
        self.template_info["remittance_info"][
            "Amount payable on recipient created invoices"] = f'''={self.stats["total_cells"][0]}'''
        self.template_info["remittance_info"][
            "Less amount owing on Goods & Services invoice"] = f'''={self.stats["total_cells"][1]}'''
        self.template_info["remittance_info"][
            "Less amount for RCIT"] = f'''={self.stats["total_cells"][2]}'''
        self.template_info["remittance_info"][
            "Less Amount for Hire Vehicle/Administrative Fixed Fee"] = f'''={self.stats["total_cells"][3]}'''
        self.template_info["remittance_info"][
            "Less Amount for Mob Hubb Fixed Fee"] = f'''={self.stats["total_cells"][4]}'''

        self.write_section("remittance_info", extra_space=True)

    def write_remittance_info_summary(self):
        self.template_info["remittance_info"][
            "Amount payable on recipient created invoices"] = f'''={self.stats["total_cells_summary"][0]}'''
        self.template_info["remittance_info"][
            "Less amount owing on Goods & Services invoice"] = f'''={self.main_sheet_title}!{self.stats["total_cells"][1]}'''
        self.template_info["remittance_info"][
            "Less amount for RCIT"] = f'''={self.main_sheet_title}!{self.stats["total_cells"][2]}'''
        self.template_info["remittance_info"][
            "Less Amount for Hire Vehicle/ Administrative Fixed Fee"] = f'''={self.main_sheet_title}!{self.stats["total_cells"][3]}'''

        self.write_section("remittance_info", extra_space=True)

    def write_section(self, section_key, source='main', extra_space=False):
        self.add_empty_lines(2)

        space = not extra_space

        if self.template_info[section_key].items():
            self.write_section_heading(section_key, space=space)

        total_start_row = self.current_row if not extra_space else self.current_row + 1
        gst_start_row = self.current_row if not extra_space else self.current_row + 1

        for info_row, val in self.template_info[section_key].items():
            if extra_space:
                self.add_empty_line()

            # Write text
            coord = f"A{self.current_row}"
            self.ws[coord] = info_row
            self.apply_formatting(
                coord, self.columns[1], ignore_num_f=True, ignore_align=True)

            # Write value
            coord = f'''{self.col_map["Amount"]}{self.current_row}'''

            if "Add GST" in info_row:
                val = f'''=0.1*SUM({self.col_map["Amount"]}{gst_start_row}:{self.col_map["Amount"]}{self.current_row - 1})'''
                self.apply_formatting(coord, "Amount")
                self.add_empty_line()

            elif "Total Toll" in info_row:
                # Can't include toll in total sums so reinitialise row that starts the section
                self.apply_formatting(coord, "Amount")
                # Add toll cell to totals information
                self.total_toll_coord = f'''{self.col_map['Amount']}{self.current_row}'''
                total_start_row = self.current_row

            elif info_row == "Total freight charges":
                gst_start_row = self.current_row
                self.apply_formatting(coord, "Amount")

                if not self.subtotal_row:
                    self.subtotal_row = self.current_row

            elif "Total charges" in info_row:
                val = f'''=SUM({self.col_map["Amount"]}{total_start_row}:{self.col_map["Amount"]}{self.current_row - 1})'''
                total_start_row = self.current_row
                gst_start_row = self.current_row
                self.apply_formatting(coord, "Amount")

            elif "invoice including GST" in info_row:
                val = f'''=SUM({self.col_map["Amount"]}{total_start_row}:{self.col_map["Amount"]}{self.current_row - 1})'''
                self.apply_formatting(coord, "Amount", bold=True)

                if source == "main":
                    self.stats["total_cells"].append(coord)

                if source == "summary":
                    self.stats["total_cells_summary"].append(coord)

            elif "Amount Remitted" in info_row:
                total_payable_cell = f'''{self.col_map["Amount"]}{total_start_row}'''
                total_deduction_range = f'''{self.col_map["Amount"]}{total_start_row + 1}:{self.col_map["Amount"]}{self.current_row - 1}'''
                val = f'''= {total_payable_cell} - SUM({total_deduction_range})'''
                self.apply_formatting(coord, "Amount", bold=True)

            else:
                self.apply_formatting(coord, "Amount")

            self.ws[coord] = val
            self.goto_next_row()

    def write_section_heading(self, key, space=True):
        possible_header = key.replace("_info", "") + "_header"

        # Write section header, if it exists in self.template_info
        if possible_header in self.template_info:
            coord = f"B{self.current_row}"
            self.ws[coord] = self.template_info[possible_header]
            self.apply_formatting(
                coord, self.columns[1], ignore_num_f=True, ignore_align=True)

            self.goto_next_row()

            if space:
                self.add_empty_line()

    def add_empty_lines(self, count):
        self.goto_next_row(increment=count)

    def add_empty_line(self):
        self.goto_next_row()

    def goto_next_row(self, increment=1):
        self.current_row += increment

    def goto_prev_row(self, decrement=1):
        self.current_row -= decrement

    # Write column header
    def write_header(self):
        self.write_list_to_row(self.current_row, self.columns)
        self.goto_next_row()

    # Write data rows to wb (i.e., rows with specific column values in self.column_values)
    # Record the first row numbers with new date values
    def write_data_rows(self):
        new_date_rows = []
        subtotal_sepr_rows = []
        prev_date = None

        for row_dict in self.column_values:
            # If no date has been encountered yet or a new date is encountered
            if row_dict["Date"] and (not prev_date or prev_date != row_dict["Date"]):
                self.stats["total_tolls_removed_by_date"][row_dict["Date"]] = 0
                self.stats["total_tolls_amount_by_date"][row_dict["Date"]] = 0
                self.stats["total_amounts_by_date"][row_dict["Date"]] = 0
                self.stats["total_jobs_by_date"][row_dict["Date"]] = 0
                prev_date = row_dict["Date"]
                self.stats["dates"].append(prev_date)
                new_date_rows.append(self.current_row)
            # If the column before 'Amount' is marked by subtotal separator
            elif row_dict[self.converter.get_col_bef("Amount")] == self.converter.get_subtotal_seperator():
                # If a separator without a new date is detected, it's the end of table
                if not prev_date:
                    break

                subtotal_sepr_rows.append(self.current_row)
                self.stats["after_date_toll_coords"].append(
                    f"{self.col_map['Amount']}{self.current_row - 1}")
                # Remove previous date once the corresponding separator is 'appended'
                prev_date = None

            # If rate includes one of 'tolls', don't write row to excel file
            if row_dict["Rate"] in self.person_info["tolls"]:
                self.stats["total_tolls_removed"] += 1
                self.stats["total_tolls_removed_by_date"][prev_date] += 1

                self.stats["total_tolls_amount"] += float(row_dict["Amount"])
                self.stats["total_tolls_amount_by_date"][prev_date] += float(
                    row_dict["Amount"])

                if row_dict["Rate"] in self.person_info["tolls_add"]:
                    add_toll_entry = f'''{row_dict["Rate"]}_tolls_removed'''
                    self.stats[add_toll_entry] = self.stats[add_toll_entry] + \
                        1 if add_toll_entry in self.stats else 1
            else:
                if prev_date and row_dict["Amount"].strip():
                    self.stats["total_jobs_by_date"][prev_date] += 1

                if not prev_date and row_dict[
                        self.converter.get_col_bef("Amount")] == self.converter.after_date_toll_text:
                    # Get rid of the last unneeded after date toll entry
                    pass
                else:
                    self.write_dict_to_row(self.current_row, row_dict)
                    self.goto_next_row()

        # Write subtotals
        self.write_subtotals(new_date_rows, subtotal_sepr_rows)

    def write_dict_to_row(self, row, row_dict):
        # Write row values using dict with column names as keys
        for col, value in row_dict.items():
            coord = f"{self.col_map[col]}{row}"

            # Cast numeric values to float
            try:
                value = float(value)
            except ValueError:
                value = value.encode("utf-8")

            # Write a percent of the original amount price
            if value and col == "Amount":
                value = self.amount_percentage * value

            self.ws[coord] = value
            self.apply_formatting(coord, col)

    def write_list_to_row(self, row, values_list, bold_l=False):
        for i, col in enumerate(self.columns):
            coord = f"{self.col_map[col]}{row}"
            self.ws[coord] = values_list[i]

            self.apply_formatting(coord, col, ignore_num_f=True, bold=bold_l)

    def write_subtotals(self, new_date_rows, subtotal_sepr_rows):
        date_index = 0

        self.stats["subtotal_coords"] = []
        for row in subtotal_sepr_rows:
            prev_date_row = new_date_rows[date_index]
            prev_date = self.stats["dates"][date_index]
            # Subtotal row is one after the subtotal seperator
            subtotal_row = row + 1

            amount_col = self.col_map["Amount"]
            # Subtotal from previous date amount to row before separator
            subtotal_val = f'''=SUM({amount_col}{prev_date_row}:{amount_col}{row - 1})'''

            coord = f'''{amount_col}{subtotal_row}'''
            self.ws[coord] = subtotal_val
            self.apply_formatting(coord, amount_col, bold=True)

            self.stats["subtotal_coords"].append(
                f'''{amount_col}{subtotal_row}''')

            # Add all values from last date to before subtotal
            for row in range(prev_date_row, subtotal_row - 1):
                amount_val = self.ws[f'''{self.col_map["Amount"]}{row}'''].value
                self.amounts.append(amount_val)
                self.stats["total_amounts_by_date"][prev_date] += amount_val if amount_val else 0
                self.stats["total_amount"] += amount_val if amount_val else 0

            date_index += 1

    # Set custom widths for excel file columns
    def set_col_widths(self):
        if "Reference" in self.columns:
            col_widths = {"Date": 13.55,
                          "Job No": 7.36,
                          "Client": 8.09,
                          "Reference": 12.55,
                          "From": 13.27,
                          "To": 11.27,
                          "Rate": 4.82,
                          "Amount": 16.73,
                          "Cash": 5.09
                          }
        elif "Jobs" in self.columns:
            col_widths = {"Date": 32.55,
                          "Jobs": 32.36,
                          "Amount": 24.5,
                          "Cash Held": 5
                          }
        else:
            col_widths = {"Date": 14.12,
                          "Job No": 10.18,
                          "From": 13.27,
                          "To": 13.36,
                          "Rate": 12.73,
                          "Amount": 18.09
                          }

        for col in col_widths:
            self.ws.column_dimensions[self.col_map[col]
                                      ].width = col_widths[col] + 0.6

    def apply_formatting(self, coord, col, ignore_num_f=False, ignore_align=False, bold=False):
        num_format = {
            "Job No": numbers.FORMAT_NUMBER,
            "Amount": '''_-$* #,##0.000_-;-$* #,##0.000_-;_-$* "-"??_-;_-@_-'''
        }
        align_format = {
            "Job No": Alignment(horizontal="right"),
            "Jobs": Alignment(horizontal="right"),
            "Amount": Alignment(horizontal="right"),
            "Cash": Alignment(horizontal="right")
        }
        default_align = Alignment(horizontal="left")

        # Apply number format and alignment formatting if row is not part of header
        if not ignore_num_f and col in num_format:
            self.ws[coord].number_format = num_format[col]

        if not ignore_align:
            self.ws[coord].alignment = align_format[col] if col in align_format else default_align

        self.ws[coord].font = Font(
            name=self.font.name, size=self.font.size, bold=True) if bold else self.font

    def print_and_log_stats(self):
        stats_output = ["######## STATS #########",
                        f"File: {self.converter.get_filename()}\n"]
        for stat, value in self.stats.items():
            stats_output.append(f"{stat}: {value}\n")

        stats_output = "\n".join(stats_output)

        check_create_folder(self.converter.logs_folder)

        logs_file = str(pathlib.Path.cwd().joinpath(
            self.converter.logs_folder, self.converter.save_name + ".txt"))
        with open(logs_file, 'w+') as obj:
            obj.write(stats_output)

        print("")
        print(stats_output)

    def subtotal_test(self):
        print("")

        check_val = round(self.converter.subtotal_check_val, 3)
        actual_val = round(
            (self.stats["total_amount"] / self.amount_percentage) + self.stats["total_tolls_amount"], 3)

        print(f"Total Value on PDF: {self.converter.subtotal_check_val}")
        print(
            f'''Total Converted Value: ({self.stats["total_amount"]} / {self.amount_percentage}) + {self.stats["total_tolls_amount"]} = {actual_val}''')

        if check_val == actual_val:
            print(f"TEST: PASSED")
        else:
            print(f"TEST: FAILED")
            exit(1)

    # Save the workbook (i.e. file)
    def save(self):
        # Create an isolated folder for the file
        check_create_folder(self.converter.output_folder)

        # Delete the csv file
        self.converter.csv_file.unlink()

        excel_file = str(
            pathlib.Path.cwd().joinpath(self.converter.get_output_folder(), self.converter.save_name + ".xlsx"))

        try:
            self.wb.save(excel_file)
        except PermissionError:
            print("Please close the excel file and try again.")
            return

        self.print_and_log_stats()

    def add_page_break(self):
        page_break = Break(id=self.current_row)
        self.ws.row_breaks.append(page_break)

    def get_subtotal_coords(self):
        return self.stats["subtotal_coords"]

    @property
    def toll_check_col(self):
        return self.__toll_check_col

    @toll_check_col.setter
    def toll_check_col(self, value):
        self.__toll_check_col = value

    @property
    def font(self):
        return self.__font

    @font.setter
    def font(self, value):
        self.__font = value

    @property
    def subtotal_row(self):
        return self.__subtotal_row

    @subtotal_row.setter
    def subtotal_row(self, value):
        self.__subtotal_row = value
